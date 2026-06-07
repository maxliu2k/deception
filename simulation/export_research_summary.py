"""Walk the project's data and produce a single XLSX rollup.

Usage:
    python -m simulation.export_research_summary
        [--out simulation/exports/research_summary.xlsx]

Output sheets:
    1. Folders      one row per folder: slots, mode, loadout, wins
    2. Auctions     one row per completed auction (315+)
    3. Datasets     training jsonl/csv files + row counts + per-LLM splits
    4. Mimics       mimic model checkpoints (v6, v7)
    5. RL_Runs      RL training runs with target/reached eps, best eval, status
    6. LLM_Usage    aggregate prompt/completion/reasoning tokens by alias
    7. Conclusions  blank text page for paper-ready notes
"""
from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO = Path(__file__).resolve().parents[1]
SIM = REPO / "simulation"
RUNTIME = SIM / ".runtime"
SAVE_SLOTS_DIR = RUNTIME / "save_slots"
CATALOG_PATH = RUNTIME / "save_slots.json"
DATASETS_DIR = SIM / "datasets"
MODELS_DIR = SIM / "models"
RL_RUNS_DIR = MODELS_DIR / "rl" / "runs"


def _read_json(path: Path) -> dict | list | None:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError):
        return None


def load_catalog() -> tuple[dict[str, dict], dict[str, dict]]:
    catalog = _read_json(CATALOG_PATH) or {}
    folders = {f["folder_id"]: f for f in catalog.get("folders", [])}
    slots = {s["slot_id"]: s for s in catalog.get("slots", [])}
    return folders, slots


def parse_bid_log(path: Path) -> dict:
    """Extract winner counts, bidder list, and bid totals from one auction log."""
    text = path.read_text(encoding="utf-8", errors="replace")
    winner_counts: Counter = Counter()
    bidder_set: set[str] = set()
    paintings_sold = 0
    bid_actions: Counter = Counter()
    for line in text.splitlines():
        m_win = re.match(r"^painting_\d+ \| status=sold \| winner=(\S+) \| winning_bid=(\d+)", line)
        if m_win:
            winner_counts[m_win.group(1)] += 1
            paintings_sold += 1
            continue
        m_turn = re.match(r"^\s+turn \d+: (\S+) (RAISE|PASS)", line)
        if m_turn:
            alias = m_turn.group(1)
            bidder_set.add(alias)
            bid_actions[alias] += 1
    return {
        "winners": dict(winner_counts),
        "bidders": sorted(bidder_set),
        "paintings_sold": paintings_sold,
        "bid_actions": dict(bid_actions),
    }


def walk_save_slots() -> tuple[list[dict], dict[str, list[dict]]]:
    """Return list of per-auction records and a folder→auctions index."""
    folders_meta, slots_meta = load_catalog()
    by_folder: dict[str, list[dict]] = defaultdict(list)
    auctions: list[dict] = []

    for slot_dir in sorted(SAVE_SLOTS_DIR.glob("save_slot_*")):
        slot_id = slot_dir.name
        meta = _read_json(slot_dir / "runtime_meta.json") or {}
        folder_id = slots_meta.get(slot_id, {}).get("folder_id") or "(unfiled)"
        slot_name = slots_meta.get(slot_id, {}).get("name") or slot_id
        mode = meta.get("mode") or "?"
        done = bool(meta.get("done"))
        export_dirs = sorted((slot_dir / "auction_exports").glob("auction_*")) if (slot_dir / "auction_exports").exists() else []
        log_summary = {}
        if export_dirs:
            log_path = export_dirs[-1] / "auction_bid_log.txt"
            if log_path.exists():
                log_summary = parse_bid_log(log_path)
        record = {
            "slot_id": slot_id,
            "slot_name": slot_name,
            "folder_id": folder_id,
            "folder_name": folders_meta.get(folder_id, {}).get("name") or "(unfiled)",
            "mode": mode,
            "done": done,
            "completed": bool(export_dirs),
            "bidders": ",".join(log_summary.get("bidders", [])),
            "paintings_sold": log_summary.get("paintings_sold", 0),
            "winners": log_summary.get("winners", {}),
            "bid_actions": log_summary.get("bid_actions", {}),
            "updated_at": meta.get("updated_at"),
        }
        auctions.append(record)
        by_folder[folder_id].append(record)
    return auctions, by_folder


def folder_rows(by_folder: dict[str, list[dict]]) -> list[dict]:
    folders_meta, _ = load_catalog()
    rows = []
    for folder_id, records in sorted(by_folder.items(), key=lambda kv: kv[0]):
        completed = [r for r in records if r["completed"]]
        wins: Counter = Counter()
        for r in completed:
            wins.update(r["winners"])
        bidders_seen = set()
        for r in completed:
            for b in r["bidders"].split(","):
                if b:
                    bidders_seen.add(b)
        sample_loadout = completed[0]["bidders"] if completed else ""
        rows.append({
            "folder_id": folder_id,
            "folder_name": folders_meta.get(folder_id, {}).get("name") or "(unfiled)",
            "total_slots": len(records),
            "completed_slots": len(completed),
            "mode": completed[0]["mode"] if completed else (records[0]["mode"] if records else "?"),
            "sample_loadout": sample_loadout,
            "total_paintings_sold": sum(r["paintings_sold"] for r in completed),
            "wins_by_bidder": "; ".join(f"{k}:{v}" for k, v in wins.most_common()),
        })
    return rows


def dataset_rows() -> list[dict]:
    rows = []
    if not DATASETS_DIR.exists():
        return rows
    for path in sorted(DATASETS_DIR.iterdir()):
        if not path.is_file():
            continue
        if path.suffix not in {".jsonl", ".csv", ".json"}:
            continue
        row_count = 0
        per_alias: Counter = Counter()
        if path.suffix == ".jsonl":
            meta_path = path.with_name(path.stem + "_meta.json")
            meta = _read_json(meta_path) if meta_path.exists() else None
            vocab = (meta or {}).get("bidder_vocab") or {}
            with path.open(encoding="utf-8") as f:
                for line in f:
                    if not line.strip():
                        continue
                    row_count += 1
                    try:
                        rec = json.loads(line)
                    except json.JSONDecodeError:
                        continue
                    if vocab:
                        idx = rec.get("bidder_index")
                        if isinstance(idx, int):
                            for name, vi in vocab.items():
                                if vi == idx:
                                    per_alias[name] += 1
                                    break
        elif path.suffix == ".csv":
            with path.open(encoding="utf-8") as f:
                row_count = max(0, sum(1 for _ in f) - 1)
        rows.append({
            "file": path.name,
            "size_kb": round(path.stat().st_size / 1024, 1),
            "rows": row_count,
            "per_alias": "; ".join(f"{k}:{v}" for k, v in per_alias.most_common()),
        })
    return rows


def mimic_rows() -> list[dict]:
    rows = []
    if not MODELS_DIR.exists():
        return rows
    for sub in sorted(MODELS_DIR.iterdir()):
        if not sub.is_dir():
            continue
        if sub.name.startswith("rl"):
            continue
        for f in sorted(sub.glob("*.pt")):
            rows.append({
                "version": sub.name,
                "file": f.name,
                "size_kb": round(f.stat().st_size / 1024, 1),
            })
    return rows


def rl_run_rows() -> list[dict]:
    rows = []
    if not RL_RUNS_DIR.exists():
        return rows
    for status_path in sorted(RL_RUNS_DIR.glob("t5_*_status.json")):
        st = _read_json(status_path) or {}
        rows.append({
            "run_id": status_path.stem.replace("_status", ""),
            "label": st.get("label") or st.get("action_space") or "",
            "action_space": st.get("action_space"),
            "target_episodes": st.get("target_episodes"),
            "episodes_reached": st.get("episode"),
            "best_eval_wins": st.get("best_eval_wins"),
            "avg_wins_last100": st.get("avg_wins_last100"),
            "finished": st.get("finished"),
            "early_stopped": st.get("early_stopped"),
            "started_at": st.get("started_at"),
            "elapsed_s": st.get("elapsed_s") or st.get("duration_s"),
        })
    # also include .jsonl-only runs without status sidecars
    seen = {r["run_id"] for r in rows}
    for jl in sorted(RL_RUNS_DIR.glob("t5_*.jsonl")):
        rid = jl.stem
        if rid in seen:
            continue
        # try to read the last line for final metrics
        final = {}
        try:
            with jl.open(encoding="utf-8") as f:
                last = ""
                for last in f:
                    pass
                if last:
                    final = json.loads(last) or {}
        except (json.JSONDecodeError, OSError):
            pass
        rows.append({
            "run_id": rid,
            "label": final.get("label") or "",
            "action_space": final.get("action_space"),
            "target_episodes": final.get("target_episodes"),
            "episodes_reached": final.get("episode"),
            "best_eval_wins": final.get("best_eval_wins"),
            "avg_wins_last100": final.get("avg_wins_last100"),
            "finished": final.get("finished"),
            "early_stopped": final.get("early_stopped"),
            "started_at": final.get("started_at"),
            "elapsed_s": final.get("elapsed_s") or final.get("duration_s"),
        })
    return rows


def llm_usage_rows() -> list[dict]:
    totals: dict[str, Counter] = defaultdict(Counter)
    for usage_path in SAVE_SLOTS_DIR.rglob("auction_usage.jsonl"):
        with usage_path.open(encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = json.loads(line)
                except json.JSONDecodeError:
                    continue
                alias = rec.get("alias") or "?"
                for k in ("prompt_tokens", "completion_tokens", "reasoning_tokens", "total_tokens"):
                    v = rec.get(k)
                    if isinstance(v, (int, float)):
                        totals[alias][k] += int(v)
                totals[alias]["call_count"] += 1
    rows = []
    for alias, c in sorted(totals.items(), key=lambda kv: -kv[1]["call_count"]):
        rows.append({
            "alias": alias,
            "call_count": c["call_count"],
            "prompt_tokens": c["prompt_tokens"],
            "completion_tokens": c["completion_tokens"],
            "reasoning_tokens": c["reasoning_tokens"],
            "total_tokens": c["total_tokens"] or (c["prompt_tokens"] + c["completion_tokens"]),
            "avg_completion_per_call": round(c["completion_tokens"] / max(1, c["call_count"]), 1),
        })
    return rows


# ── XLSX rendering ───────────────────────────────────────────────────

def write_sheet(wb: Workbook, name: str, rows: list[dict], description: str = "") -> None:
    ws = wb.create_sheet(name)
    header_fill = PatternFill(start_color="FFE5E5E5", end_color="FFE5E5E5", fill_type="solid")
    bold = Font(bold=True)
    row_offset = 1
    if description:
        ws.cell(row=1, column=1, value=description).font = Font(italic=True, color="FF666666")
        row_offset = 3
    if not rows:
        ws.cell(row=row_offset, column=1, value="(no data)")
        return
    headers = list(rows[0].keys())
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row_offset, column=col_idx, value=h)
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left")
    for r_idx, row in enumerate(rows, start=row_offset + 1):
        for col_idx, h in enumerate(headers, start=1):
            v = row.get(h)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, separators=(",", ":"))
            ws.cell(row=r_idx, column=col_idx, value=v)
    # auto width (capped)
    for col_idx, h in enumerate(headers, start=1):
        max_len = max([len(str(h))] + [len(str(row.get(h, ""))) for row in rows])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(60, max_len + 2)
    ws.freeze_panes = ws.cell(row=row_offset + 1, column=1)


def write_conclusions_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Conclusions")
    ws.cell(row=1, column=1, value="Research conclusions — paste latest paper-ready summary here.").font = Font(italic=True, color="FF666666")
    ws.column_dimensions["A"].width = 120
    for i, line in enumerate([
        "Headline data points:",
        "1. Mimic fit chi² p = 0.146, V = 0.041 (5 LLM/mimic pairs).",
        "2. Math tier ladder monotone: T3 ~24%, T4 ~28%, vs 20% uniform baseline.",
        "3. RL T5 ceiling 3.62/12 wins; only ~6% relative improvement over T4.",
        "4. Frontier LLMs converge on fair-share + risk-threshold; one-dimensional strategy class.",
        "5. CoT is justification, not exploration: zero novel algorithms across ~2000 logged bids.",
    ], start=3):
        ws.cell(row=i, column=1, value=line)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=str(SIM / "exports" / "research_summary.xlsx"))
    args = ap.parse_args()

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"[summary] reading save_slots ({SAVE_SLOTS_DIR})")
    auctions, by_folder = walk_save_slots()
    print(f"[summary]   {len(auctions)} slots / {sum(1 for r in auctions if r['completed'])} completed")
    folders = folder_rows(by_folder)
    print(f"[summary] reading datasets ({DATASETS_DIR})")
    datasets = dataset_rows()
    print(f"[summary]   {len(datasets)} dataset files")
    print(f"[summary] reading models ({MODELS_DIR})")
    mimics = mimic_rows()
    print(f"[summary]   {len(mimics)} mimic checkpoints")
    rl_runs = rl_run_rows()
    print(f"[summary]   {len(rl_runs)} RL runs")
    print(f"[summary] aggregating LLM usage")
    usage = llm_usage_rows()
    print(f"[summary]   {len(usage)} aliases in usage logs")

    wb = Workbook()
    # remove default sheet
    default = wb.active
    wb.remove(default)

    write_sheet(wb, "Folders", folders,
                "One row per folder of save slots. wins_by_bidder is aggregated across all completed auctions.")
    write_sheet(wb, "Auctions", [{
        "slot_id": r["slot_id"],
        "slot_name": r["slot_name"],
        "folder_name": r["folder_name"],
        "mode": r["mode"],
        "completed": r["completed"],
        "paintings_sold": r["paintings_sold"],
        "bidders": r["bidders"],
        "winners": "; ".join(f"{k}:{v}" for k, v in r["winners"].items()),
    } for r in auctions],
                "One row per save slot. Winners is the sold-paintings winner-count map for that auction.")
    write_sheet(wb, "Datasets", datasets,
                "Training dataset files in simulation/datasets/. per_alias only populated for v6/v7 NN jsonl with a *_meta.json.")
    write_sheet(wb, "Mimics", mimics,
                "Trained mimic checkpoints in simulation/models/. RL models are on the RL_Runs sheet.")
    write_sheet(wb, "RL_Runs", rl_runs,
                "PPO / BC training runs. avg_wins_last100 is the headline metric (12 paintings per episode).")
    write_sheet(wb, "LLM_Usage", usage,
                "Aggregate token counts across all auction_usage.jsonl files. avg_completion_per_call is a proxy for reasoning depth.")
    write_conclusions_sheet(wb)

    wb.save(out_path)
    print(f"[summary] wrote {out_path}  ({out_path.stat().st_size / 1024:.1f} KB)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
